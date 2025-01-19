import pandas as pd
import json
import os
import openpyxl
import re

def extract_guid_from_url(url):
    """URL에서 _0/ 다음에 오는 GUID를 추출하는 함수"""
    if url:
        try:
            # _0/ 다음에 오는 GUID를 찾는 패턴
            pattern = r'_0/([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})'
            match = re.search(pattern, url.lower())
            if match:
                return match.group(1)  # GUID 부분만 반환
        except Exception as e:
            print(f"GUID 추출 중 오류 발생: {str(e)}")
    return None

def get_hyperlink_value(excel_path, cell):
    """셀의 하이퍼링크 URL을 가져와서 GUID와 전체 URL을 반환하는 함수"""
    try:
        if cell.hyperlink:
            url = cell.hyperlink.target
            guid = extract_guid_from_url(url)
            if guid:
                return {"guid": guid, "url": url}
            else:
                print(f"GUID를 찾을 수 없음: {url}")
        return None
    except Exception as e:
        print(f"하이퍼링크 처리 중 오류 발생: {str(e)}")
        return None

def excel_to_json(file_path):
    try:
        # 워크북 직접 로드
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Assigned Workflow']
        
        # 데이터를 저장할 리스트
        records = []
        
        # 9행부터 데이터 읽기
        row_num = 9
        while True:
            # B열의 값 확인
            fsli = ws.cell(row=row_num, column=2).value
            
            # 데이터가 없으면 종료
            if not fsli:
                break
                
            record = {
                'FSLI(s)': ws.cell(row=row_num, column=2).value,
                'Primary FSLI': ws.cell(row=row_num, column=3).value,
                'EGA': ws.cell(row=row_num, column=4).value,
                'Type': ws.cell(row=row_num, column=5).value,
            }
            
            # 하이퍼링크 정보 가져오기
            link_info = get_hyperlink_value(file_path, ws.cell(row=row_num, column=27))
            if link_info:
                record['Link'] = link_info['guid']  # 기존의 Link 필드는 guid 유지
                record['hyperlink'] = link_info['url']  # 새로운 hyperlink 필드 추가
            
            records.append(record)
            row_num += 1
        
        wb.close()
        
        # JSON 파일 저장 경로를 json 폴더로 설정
        output_dir = os.path.join(os.getcwd(), 'json')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        output_path = os.path.join(output_dir, 'output.json')
        
        # JSON 파일로 저장
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        
        print(f"JSON 파일이 다음 경로에 생성되었습니다: {output_path}")
        
        return json.dumps(records, ensure_ascii=False, indent=2)
        
    except Exception as e:
        print(f"에러가 발생했습니다: {str(e)}")
        print(f"에러 타입: {type(e)}")
        return None

if __name__ == "__main__":
    file_path = r"C:\Users\jkim564\OneDrive - PwC\Download\Assigned Workflow by FSLI - 루미레즈코리아2412 (2025-01-19T16-41-23).XLSX"

    json_data = excel_to_json(file_path)
    if json_data:
        print("\n변환된 JSON 데이터 샘플 (처음 2개 항목):")
        json_obj = json.loads(json_data)
        print(json.dumps(json_obj[:2], ensure_ascii=False, indent=2))
        print(f"\n총 {len(json_obj)}개의 항목이 변환되었습니다.")