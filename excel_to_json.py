from openpyxl import load_workbook
import json
from typing import Dict, List, Any, Optional
from pathlib import Path
import io
from datetime import datetime
import os




class ExcelDocumentParser:
    def __init__(self, file_content: bytes, file_name: str = ""):
        """Excel 문서 파서 초기화"""
        self.wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        # 보이는 시트만 필터링
        self.visible_sheets = []
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            if sheet.sheet_state == 'visible':
                self.visible_sheets.append(sheet_name)
        
        # 기본 메타데이터 구성
        self.document_structure = {
            'metadata': {
                'file_name': file_name,
                'total_sheets': len(self.visible_sheets),
                'sheet_names': self.visible_sheets,
                'sheets_info': {},
                'last_modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            },
            'sheets': {}
        }

    def _get_sheet_info(self, sheet) -> Dict:
        """시트의 기본 정보 추출"""
        return {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'has_merged_cells': bool(sheet.merged_cells),
            'sheet_state': sheet.sheet_state  # 시트 상태 추가
        }

    def _get_cell_value(self, cell: Any) -> Optional[Any]:
        """셀 값을 적절한 형태로 반환"""
        if cell.value is None:
            return None
        
        value = cell.value
        try:
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d')
            return value
        except:
            return str(value)

    def parse_sheet(self, sheet) -> Dict:
        """시트 내용 파싱"""
        sheet_content = []
        current_row_index = 0

        for row in sheet.rows:
            row_content = {}
            has_value = False

            for cell in row:
                value = self._get_cell_value(cell)
                if value is not None:  # 빈 셀 제외
                    has_value = True
                    row_content[cell.column_letter] = {
                        'value': value,
                        'coordinate': cell.coordinate
                    }
            
            if has_value:  # 값이 있는 행만 추가
                sheet_content.append({
                    'row_index': current_row_index,
                    'content': row_content
                })
            
            current_row_index += 1

        return sheet_content

    def parse_document(self) -> Dict:
        """전체 문서 파싱 (숨겨진 시트 제외)"""
        for sheet_name in self.visible_sheets:
            sheet = self.wb[sheet_name]
            
            # 시트 정보 저장
            self.document_structure['metadata']['sheets_info'][sheet_name] = self._get_sheet_info(sheet)
            
            # 시트 내용 파싱
            sheet_content = self.parse_sheet(sheet)
            if sheet_content:  # 내용이 있는 시트만 추가
                self.document_structure['sheets'][sheet_name] = sheet_content

        return self.document_structure



def get_xlsm_files(folder_path: str) -> list:
    """지정된 폴더에서 .xlsm 파일만 가져오기"""
    folder = Path(folder_path)
    return [file for file in folder.glob("*.xlsm") if file.is_file()]


def process_xlsm_file(file_path: Path) -> Dict:
    """.xlsm 파일을 ExcelDocumentParser를 이용해 JSON 형식으로 변환"""
    try:
        with file_path.open("rb") as file:
            file_content = file.read()
            parser = ExcelDocumentParser(file_content, file_path.name)
            return parser.parse_document()

    except Exception as e:
        print(f"Error processing file {file_path.name}: {e}")
        return {}

def convert_xlsm_folder_to_json(folder_path: str):
    """폴더 내 .xlsm 파일들을 JSON으로 변환 후 저장"""
    xlsm_files = get_xlsm_files(folder_path)

    output_folder = os.getcwd() + r"\json"
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    for xlsm_file in xlsm_files:
        result = process_xlsm_file(xlsm_file)
        if result:
            output_file = output_path / f"{xlsm_file.stem}.json"
            with output_file.open("w", encoding="utf-8") as json_file:
                json.dump(result, json_file, ensure_ascii=False, indent=4)


# 경로 생성기
def making_excel_path(engagement_id):
    current_dir = os.getcwd()
    count_slash = 0
    for i, alpha in enumerate(current_dir):
        if(alpha == "\\"):
            index_slash = i
            count_slash += 1
        if (count_slash == 3):
            break
    
    start_path = current_dir[:index_slash+1]
    mid_path = r"AppData\Local\AuraOffline\kor\Documents\\"
    end_path = engagement_id
    path = start_path + mid_path + end_path
    return path


def extract_engagement_id(url: str) -> str:
    """Aura URL에서 engagement ID를 추출 (#/ 다음부터 그 다음 / 전까지의 값)"""
    try:
        start = url.find('#/') + 2  # '#/' 다음 위치
        if start > 1:  # '#/'를 찾은 경우
            remaining = url[start:]  # '#/' 이후의 문자열
            end = remaining.find('/')  # 다음 '/'의 위치
            if end > 0:  # '/'를 찾은 경우
                return remaining[:end]
        raise ValueError("URL에서 engagement ID를 찾을 수 없습니다.")
    except Exception as e:
        print(f"Error: {str(e)}")
        return None

# 사용 예시
url = "https://kr-platinum.aura.pwcglb.com/#/519f4020-66d2-4a70-96d5-b00fc2d0e9e0/execute/execute"
engagement_id = extract_engagement_id(url)
print(engagement_id)  # 출력: 519f4020-66d2-4a70-96d5-b00fc2d0e9e0

# 기존 코드에서 engagement_id 직접 할당 대신 이 함수 사용
folder_path = making_excel_path(engagement_id)
print(folder_path)
convert_xlsm_folder_to_json(folder_path=folder_path)
